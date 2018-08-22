import os.path as opath
import os
import pandas as pd
import csv
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
#
from __path_organizer import aggr_dpath, h1data_dpath


HOURS = list(range(6, 24)) + [0, 1]
TERMINALS1 = ['T1', 'T2', 'T3', 'B', 'X']
TERMINALS2 = ['T1', 'T2', 'T3', 'T4', 'X']

cm_align=Alignment(horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
rm_align=Alignment(horizontal='right',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)


_rgb = lambda r, g, b: (r / float(255), g / float(255), b / float(255))
clists = (
    'blue', 'green', 'red', 'cyan', 'magenta', 'black',
    _rgb(255, 165, 0),  # orange
    _rgb(238, 130, 238),  # violet
    _rgb(255, 228, 225),  # misty rose
    _rgb(127, 255, 212),  # aqua-marine
    'yellow',
    _rgb(220, 220, 220),  # gray
    _rgb(255, 165, 0),  # orange
    'black'
)
mlists = (
    'o',  #    circle
    'v',  #    triangle_down
    '^',  #    triangle_up
    '<',  #    triangle_left
    '>',  #    triangle_right
    's',  #    square
    'p',  #    pentagon
    '*',  #    star
    '+',  #    plus
    'x',  #    x
    'D',  #    diamond
    'h',  #    hexagon1
    '1',  #    tri_down
    '2',  #    tri_up
    '3',  #    tri_left
    '4',  #    tri_right
    '8',  #    octagon
    'H',  #    hexagon2
    'd',  #    thin_diamond
    '|',  #    vline
    '_',  #    hline
    '.',  #    point
    ',',  #    pixel

    'D',  #    diamond
    '8',  #    octagon
    )



def gen_xlsx_file(terminals, ps_count, ofpath):
    wb = Workbook()
    ws = wb.active
    ws.title = "RawNumber"
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
    ori_cell, dest_cell = ws.cell(row=1, column=1), ws.cell(row=1, column=2)
    ori_cell.value = 'Origin'
    dest_cell.value = 'Destination'
    for cell in [ori_cell, dest_cell]:
        cell.alignment = cm_align
    for i, tn in enumerate(terminals + ['Total']):
        o_cell, d_cell = ws.cell(row=i + 3, column=1), ws.cell(row=2, column=i + 2)
        for cell in [o_cell, d_cell]:
            cell.value = tn
            cell.alignment = cm_align
    for i, prevLoc in enumerate(terminals):
        for j, startLoc in enumerate(terminals):
            k = prevLoc, startLoc
            t_cell = ws.cell(row=i + 3, column=j + 2)
            t_cell.value = ps_count[k] if k in ps_count else 0
            t_cell.alignment = rm_align
        j += 1
        t_cell = ws.cell(row=i + 3, column=j + 2)
        bIndex = t_cell.col_idx - len(terminals)
        eIndex = t_cell.col_idx - 1
        t_cell.value = "=SUM(%s%d:%s%d)" % (get_column_letter(bIndex), t_cell.row,
                                            get_column_letter(eIndex), t_cell.row)
    for i, tn in enumerate(terminals + ['Total']):
        t_cell = ws.cell(row=8, column=i + 2)
        bIndex = t_cell.row - len(terminals)
        eIndex = t_cell.row - 1
        colName = t_cell.column
        t_cell.value = "=SUM(%s%d:%s%d)" % (colName, bIndex, colName, eIndex)
        #
    ws1 = wb.copy_worksheet(ws)
    ws1.title = "Ratio"
    totalCellCol = get_column_letter(2 + len(terminals))
    for i in range(len(terminals)):
        for j in range(len(terminals)):
            cell = ws1.cell(row=i + 3, column=j + 2)
            cell.value = "=%s!%s/%s!%s%d" % (ws.title, cell.coordinate, ws.title, totalCellCol, cell.row)
            cell.number_format = '0.00%'

    totalCell = ws['G8']
    for i, tn in enumerate(terminals + ['Total']):
        cell = ws1.cell(row=8, column=i + 2)
        cell.value = "=%s!%s/%s!%s" % (ws.title, cell.coordinate, ws.title, totalCell.coordinate)
        cell.number_format = '0.00%'
        #
        cell = ws1.cell(row=i + 3, column=7)
        cell.value = "=%s!%s/%s!%s" % (ws.title, cell.coordinate, ws.title, totalCell.coordinate)
        cell.number_format = '0.00%'
    wb.save(filename=ofpath)


def arrange_datasets():
    yyyys = ['2009', '2010']
    for yyyy in yyyys:
        terminals = TERMINALS1 if yyyy != '2018' else TERMINALS2
        df = pd.read_csv(opath.join(aggr_dpath, 'apTrip-%s.csv' % yyyy))
        ydf = df.groupby(['previous_dropoff_loc', 'start_loc']).count()['taxi_id'].reset_index(name="count")
        ps_count = {}
        for prevLoc, startLoc, count in ydf.values:
            ps_count[prevLoc, startLoc] = count
        ofpath = opath.join(h1data_dpath, 'AirportFlow-%s.xlsx' % yyyy)
        gen_xlsx_file(terminals, ps_count, ofpath)
        #
        mdf = df.groupby(['month', 'previous_dropoff_loc', 'start_loc']).count()['taxi_id'].reset_index(name="count")
        mps_count = {}
        months = set()
        for month, prevLoc, startLoc, count in mdf.values:
            mps_count[month, prevLoc, startLoc] = count
            months.add(month)
        for m in months:
            ps_count = {}
            for prevLoc in terminals:
                for startLoc in terminals:
                    k = m, prevLoc, startLoc
                    ps_count[prevLoc, startLoc] = mps_count[k] if k in mps_count else 0
            ofpath = opath.join(h1data_dpath, 'AirportFlow-%sM%02d.xlsx' % (yyyy, m))
            gen_xlsx_file(terminals, ps_count, ofpath)
        #
        hdf = df.groupby(['hour', 'previous_dropoff_loc', 'start_loc']).count()['taxi_id'].reset_index(name="count")
        hps_count = {}
        hours = set()
        for hour, prevLoc, startLoc, count in hdf.values:
            hps_count[hour, prevLoc, startLoc] = count
            hours.add(hour)
        for h in hours:
            ps_count = {}
            for prevLoc in terminals:
                for startLoc in terminals:
                    k = h, prevLoc, startLoc
                    ps_count[prevLoc, startLoc] = hps_count[k] if k in hps_count else 0
            ofpath = opath.join(h1data_dpath, 'AirportFlow-%sH%02d.xlsx' % (yyyy, h))
            gen_xlsx_file(terminals, ps_count, ofpath)


FIGSIZE = (8, 6)

def draw_chart():
    years = [2009, 2010]
    df = None
    for yyyy in years:
        fpath = opath.join(aggr_dpath, 'apTrip-%d.csv' % yyyy)
        df = pd.read_csv(fpath) if df is None else df.append(pd.read_csv(fpath))
    hdf = df.groupby(['year', 'hour', 'previous_dropoff_loc', 'start_loc']).count()['taxi_id'].reset_index(name="count")
    hours = sorted(list(set(hdf['hour'])))
    yhLocs_count = {}
    for year, hour, pt, st, count in hdf.values:
        yhLocs_count[year, hour, pt, st] = count        
    #
    hyLocs_ratio = {}
    for pt in TERMINALS1:
        for h in hours:
            #
            for y in years:
                sumTrips = sum(yhLocs_count[y, h, pt, st] for st in TERMINALS1)
                for st in TERMINALS1:
                    hyLocs_ratio[h, y, pt, st] = yhLocs_count[y, h, pt, st] / float(sumTrips) * 100
                    
    for pt in TERMINALS1:
        for st in TERMINALS1:
            img_ofpath = 'FlowChange-%s-%s.pdf' % (pt, st)
            fig = plt.figure(figsize=FIGSIZE)
            ax = fig.add_subplot(111)
            ax.set_xlabel('Hour')
            ax.set_ylabel('%')
            for i, y in enumerate([2009, 2010]):
                plt.plot(range(len(hours)), [hyLocs_ratio[h, y, pt, st] if (h, y, pt, st) in hyLocs_ratio else 0 for h in HOURS],
                         color=clists[i], marker=mlists[i])
            plt.legend(['2009', '2010'], ncol=1)
            plt.xticks(range(len(HOURS)), HOURS)        
            plt.savefig(img_ofpath, bbox_inches='tight', pad_inches=0)

        
        
        
            hyLocs_ratio[6, 2009, 'T3', 'X']
        
    hdf2009 = hdf[(hdf['year'] == 2009)]
    hdf2010 = hdf[(hdf['year'] == 2010)]
    
    
        
        
        
        


if __name__ == '__main__':
    arrange_datasets()